from flask import Blueprint, request, jsonify, send_file
from datetime import datetime
from io import BytesIO
import openpyxl
from config.database import records, fs, BACKUP_DIR, statistics
from bson.objectid import ObjectId
import logging

logger = logging.getLogger(__name__)
backup_bp = Blueprint('backup', __name__, url_prefix='/api')

@backup_bp.route('/create_backup', methods=['POST'])
def create_backup():
    try:
        data = request.get_json()
        backup_name = data.get('name', '').strip() if data else ''
        
        if not backup_name:
            return jsonify({'error': '请输入备份名称'}), 400
            
        # 1. 获取当前Excel模板文件
        template = fs.find_one({'filename': 'template.xlsx'})
        if not template:
            return jsonify({'error': '未找到Excel模板文件'}), 404
            
        # 从模板创建新的工作簿
        template_data = BytesIO(template.read())
        template_wb = openpyxl.load_workbook(template_data)
        template_sheet = template_wb.active
        
        # 2. 获取所有数据库记录并按前缀分组
        db_records = list(records.find())
        prefix_groups = {}
        
        # 对数据库记录按前缀分组
        for record in db_records:
            prefix = record['number'][:6] if 'number' in record else ''
            if prefix:
                if prefix not in prefix_groups:
                    prefix_groups[prefix] = []
                prefix_groups[prefix].append(record)
        
        # 3. 处理每个前缀组
        red_font = openpyxl.styles.Font(color="FF0000")
        
        for prefix, group_records in prefix_groups.items():
            target_column = None
            target_row = None
            
            # 在Excel中查找匹配的前缀列
            for col in range(1, template_sheet.max_column + 1, 2):
                found = False
                for row in range(1, template_sheet.max_row + 1):
                    cell_value = str(template_sheet.cell(row=row, column=col).value or '')
                    if cell_value.startswith(prefix):
                        target_column = col
                        # 找到该列的最后一个非空单元格
                        last_row = row
                        while last_row < template_sheet.max_row and template_sheet.cell(row=last_row + 1, column=col).value is not None:
                            last_row += 1
                        target_row = last_row + 1
                        found = True
                        break
                if found:
                    break
            
            if not target_column:
                # 如果没找到匹配的列，创建新列
                target_column = template_sheet.max_column + 1
                target_row = 1
            
            # 写入该前缀组的所有记录
            for record in group_records:
                # 检查记录是否已存在
                exists = False
                for row in range(1, template_sheet.max_row + 1):
                    if str(template_sheet.cell(row=row, column=target_column).value) == record['number']:
                        exists = True
                        break
                
                if not exists:
                    # 写入数据并设置红色字体
                    cell1 = template_sheet.cell(row=target_row, column=target_column, value=record['number'])
                    cell1.font = red_font
                    
                    cell2 = template_sheet.cell(row=target_row, column=target_column + 1, value='扫描添加')
                    cell2.font = red_font
                    
                    target_row += 1
        
        # 4. 保存备份文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"backup_{timestamp}_{backup_name.replace(' ', '_')}.xlsx"
        
        # 保存到内存中
        excel_data = BytesIO()
        template_wb.save(excel_data)
        excel_data.seek(0)
        
        # 存储到GridFS
        file_id = fs.put(
            excel_data.getvalue(),  # 使用getvalue()而不是read()
            filename=backup_filename,
            backup_name=backup_name,
            timestamp=datetime.now(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        return jsonify({
            'message': '备份创建成功',
            'backup': {
                'id': str(file_id),
                'filename': backup_filename,
                'name': backup_name,
                'time': timestamp
            }
        })
        
    except Exception as e:
        logger.error(f"创建备份失败: {str(e)}")
        return jsonify({'error': str(e)}), 500

@backup_bp.route('/get_backups')
def get_backups():
    try:
        # 从 GridFS 获取所有备份文件
        backup_files = list(fs.find({"backup_name": {"$exists": True}}).sort('timestamp', -1))
        
        backups = []
        for file in backup_files:
            backups.append({
                'id': str(file._id),
                'name': file.backup_name,
                'filename': file.filename,
                'time': file.timestamp.strftime('%Y-%m-%d %H:%M:%S') if hasattr(file, 'timestamp') else '未知时间'
            })
        
        return jsonify(backups)
        
    except Exception as e:
        logger.error(f"获取备份列表失败: {str(e)}")
        return jsonify({'error': str(e)}), 500

@backup_bp.route('/download_backup/<file_id>')
def download_backup(file_id):
    try:
        # 获取文件
        file = fs.get(ObjectId(file_id))
        if not file:
            return jsonify({'error': '文件不存在'}), 404
            
        # 读取文件内容
        file_data = file.read()
        
        # 创建内存文件对象
        memory_file = BytesIO(file_data)
        
        # 获取文件名，移除可能的特殊字符
        filename = file.filename.replace(' ', '_')
        
        # 发送文件
        return send_file(
            memory_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"下载备份失败: {str(e)}")
        return jsonify({'error': str(e)}), 500

@backup_bp.route('/delete_backup/<file_id>', methods=['POST'])
def delete_backup(file_id):
    try:
        fs.delete(ObjectId(file_id))
        return jsonify({'message': '备份已删除'})
    except Exception as e:
        logger.error(f"删除备份失败: {str(e)}")
        return jsonify({'error': str(e)}), 500

@backup_bp.route('/clear_database', methods=['POST'])
def clear_database():
    try:
        data = request.get_json()
        password = data.get('password')
        
        # 验证密码
        if password != 'admin123':  # 这里设置您想要的管理员密码
            return jsonify({'error': '密码错误'}), 403
            
        # 清除所有集合的数据
        records_count = records.delete_many({}).deleted_count
        statistics.delete_many({})
        
        # 删除所有文件（除了模板文件）
        for file in fs.find({'filename': {'$ne': 'template.xlsx'}}):
            fs.delete(file._id)
            
        return jsonify({
            'message': '数据库清除成功',
            'count': records_count
        })
        
    except Exception as e:
        logger.error(f"清除数据库失败: {str(e)}")
        return jsonify({'error': str(e)}), 500 