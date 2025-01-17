from flask import Blueprint, request, jsonify
from datetime import datetime
from config.database import records, fs
import logging
import openpyxl
from io import BytesIO

logger = logging.getLogger(__name__)
scan_bp = Blueprint('scan', __name__, url_prefix='/api')

@scan_bp.route('/scan', methods=['POST'])
def scan():
    try:
        data = request.get_json()
        if not data or 'number' not in data:
            logger.warning("无效的请求数据")
            return jsonify({
                'message': '无效的请求数据',
                'status': 'invalid'
            })
            
        number = data['number'].strip()
        if not number:
            logger.warning("编号为空")
            return jsonify({
                'message': '编号不能为空',
                'status': 'invalid'
            })
            
        # 1. 首先验证是否为纯数字且长度大于6位
        if not (number.isdigit() and len(number) >= 6):
            logger.warning(f"无效的编号格式: {number}")
            return jsonify({
                'message': '无效的编号格式',
                'status': 'invalid'
            })
            
        # 2. 获取前6位作为前缀
        prefix = number[:6]
        
        # 3. 检查Excel中是否存在匹配的前缀
        template = fs.find_one({'filename': 'template.xlsx'})
        if not template:
            return jsonify({
                'message': '未找到Excel模板文件',
                'status': 'invalid'
            })
            
        # 读取Excel文件并查找前缀
        workbook = openpyxl.load_workbook(BytesIO(template.read()))
        sheet = workbook.active
        
        # 查找匹配的前缀列
        target_column = None
        target_row = None
        
        # 遍历Excel列（每隔一列查找，因为数据在奇数列）
        for col in range(1, sheet.max_column + 1, 2):
            found = False
            for row in range(1, sheet.max_row + 1):
                cell_value = str(sheet.cell(row=row, column=col).value or '')
                if cell_value.startswith(prefix):
                    target_column = col
                    # 找到该列的最后一个非空单元格
                    last_row = row
                    while last_row < sheet.max_row and sheet.cell(row=last_row + 1, column=col).value is not None:
                        last_row += 1
                    target_row = last_row + 1
                    found = True
                    break
            if found:
                break
        
        # 如果没找到匹配的前缀列，返回无效
        if not target_column:
            logger.warning(f"未找到匹配的前缀: {prefix}")
            return jsonify({
                'message': '无效的编号前缀',
                'status': 'invalid'
            })
            
        # 4. 检查是否已存在于数据库
        existing_record = records.find_one({'number': number})
        if existing_record:
            logger.info(f"发现重复编号: {number}")
            return jsonify({
                'message': '此编号已存在',
                'status': 'duplicate'
            })
            
        # 5. 所有验证通过，写入数据
        red_font = openpyxl.styles.Font(color="FF0000")
        
        # 写入编号
        cell1 = sheet.cell(row=target_row, column=target_column, value=number)
        cell1.font = red_font
        
        # 写入状态
        cell2 = sheet.cell(row=target_row, column=target_column + 1, value='扫描添加')
        cell2.font = red_font
        
        # 保存Excel文件
        excel_data = BytesIO()
        workbook.save(excel_data)
        excel_data.seek(0)
        
        # 更新GridFS中的文件
        fs.delete(template._id)
        fs.put(
            excel_data.getvalue(),
            filename='template.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # 添加记录到数据库
        records.insert_one({
            'number': number,
            'prefix': prefix,
            'timestamp': datetime.now(),
            'status': 'success'
        })
        
        logger.info(f"扫描成功: {number}")
        return jsonify({
            'message': '扫描成功',
            'status': 'success'
        })
        
    except Exception as e:
        logger.error(f"扫描处理失败: {str(e)}")
        return jsonify({
            'message': f'扫描处理失败: {str(e)}',
            'status': 'invalid'
        })