from flask import Blueprint, request, jsonify, send_file
from werkzeug.utils import secure_filename
from config.database import allowed_file, fs, records
from bson.objectid import ObjectId
from io import BytesIO
import logging
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from datetime import datetime

logger = logging.getLogger(__name__)
file_bp = Blueprint('file', __name__, url_prefix='/api')

@file_bp.route('/upload_template', methods=['POST'])
def upload_template():
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有文件被上传'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400
            
        if file and allowed_file(file.filename):
            # 先删除旧的模板文件（如果存在）
            old_template = fs.find_one({'filename': 'template.xlsx'})
            if old_template:
                fs.delete(old_template._id)
            
            # 保存新的模板文件
            file_id = fs.put(
                file.read(),
                filename='template.xlsx',
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            logger.info(f"模板文件已上传: {file_id}")
            return jsonify({'message': '文件上传成功'})
        else:
            return jsonify({'error': '不支持的文件类型'}), 400
            
    except Exception as e:
        logger.error(f"上传文件失败: {str(e)}")
        return jsonify({'error': str(e)}), 500

@file_bp.route('/check_template')
def check_template():
    try:
        template = fs.find_one({'filename': 'template.xlsx'})
        return jsonify({'exists': template is not None})
    except Exception as e:
        logger.error(f"检查模板文件失败: {str(e)}")
        return jsonify({'error': str(e)}), 500

@file_bp.route('/download_template', methods=['GET'])
def download_template():
    try:
        # 获取下载类型参数
        download_type = request.args.get('type', 'current')  # current: 当前文件, all: 所有数据
        
        if download_type == 'current':
            # 下载当前Excel文件
            template = fs.find_one({'filename': 'template.xlsx'})
            if not template:
                return jsonify({'error': '未找到Excel文件'}), 404
                
            return send_file(
                BytesIO(template.read()),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'扫描数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
            
        elif download_type == 'all':
            workbook = Workbook()
            sheet = workbook.active
            
            # 创建红色字体样式
            red_font = Font(color="FF0000")  # 提前创建红色字体样式
            
            # 1. 先复制Excel文件中的数据和样式
            template = fs.find_one({'filename': 'template.xlsx'})
            if template:
                template_wb = load_workbook(BytesIO(template.read()))
                template_sheet = template_wb.active
                
                # 复制Excel中的数据和样式
                for row_idx, row in enumerate(template_sheet.iter_rows(), 1):
                    for col_idx, cell in enumerate(row, 1):
                        new_cell = sheet.cell(row=row_idx, column=col_idx, value=cell.value)
                        if cell.font and hasattr(cell.font, 'color') and cell.font.color:
                            # 如果原单元格是红色，保持红色
                            if cell.font.color.rgb == "FFFF0000":
                                new_cell.font = red_font
            
            # 2. 获取数据库中的所有记录并按前缀分组
            db_records = list(records.find())
            prefix_groups = {}
            
            # 对数据库记录按前缀分组
            for record in db_records:
                prefix = record['number'][:6]
                if prefix not in prefix_groups:
                    prefix_groups[prefix] = []
                prefix_groups[prefix].append(record)
            
            # 3. 处理每个前缀组
            for prefix, group_records in prefix_groups.items():
                target_column = None
                target_row = None
                
                # 在Excel中查找匹配的前缀列
                for col in range(1, sheet.max_column + 1, 2):
                    found = False
                    for row in range(1, sheet.max_row + 1):
                        cell_value = str(sheet.cell(row=row, column=col).value or '')
                        if cell_value.startswith(prefix):
                            target_column = col
                            # 找到该列的最后一个非空单元格
                            last_row = row
                            while sheet.cell(row=last_row + 1, column=col).value is not None:
                                last_row += 1
                            target_row = last_row + 1
                            found = True
                            break
                    if found:
                        break
                
                if not target_column:
                    # 如果没找到匹配的列，创建新列
                    target_column = sheet.max_column + 1
                    target_row = 1
                
                # 写入该前缀组的所有记录
                for record in group_records:
                    # 检查记录是否已存在
                    exists = False
                    for row in range(1, sheet.max_row + 1):
                        if str(sheet.cell(row=row, column=target_column).value) == record['number']:
                            exists = True
                            break
                    
                    if not exists:
                        # 写入数据并设置红色字体
                        cell1 = sheet.cell(row=target_row, column=target_column, value=record['number'])
                        cell1.font = red_font  # 使用红色字体
                        
                        cell2 = sheet.cell(row=target_row, column=target_column + 1, value='扫描添加')
                        cell2.font = red_font  # 使用红色字体
                        
                        target_row += 1
            
            # 保存并返回文件
            excel_data = BytesIO()
            workbook.save(excel_data)
            excel_data.seek(0)
            
            return send_file(
                excel_data,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'全部数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
            
        else:
            return jsonify({'error': '无效的下载类型'}), 400
            
    except Exception as e:
        logger.error(f"下载文件失败: {str(e)}")
        return jsonify({'error': str(e)}), 500

@file_bp.route('/delete_template', methods=['POST'])
def delete_template():
    try:
        template = fs.find_one({'filename': 'template.xlsx'})
        if template:
            fs.delete(template._id)
            logger.info("模板文件已删除")
            return jsonify({'message': '模板文件已删除'})
        else:
            return jsonify({'error': '模板文件不存在'}), 404
    except Exception as e:
        logger.error(f"删除模板文件失败: {str(e)}")
        return jsonify({'error': str(e)}), 500 