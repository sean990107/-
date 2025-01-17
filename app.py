from flask import Flask, render_template, request, jsonify, send_file
import csv
import os
from werkzeug.utils import secure_filename
from flask_cors import CORS
import logging
import openpyxl  # 使用 openpyxl 处理 Excel 文件
import boto3  # 添加 S3 存储支持
from datetime import datetime
import shutil

# 配置日志
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 限制上传文件大小为16MB
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# S3 配置
S3_BUCKET = os.environ.get('S3_BUCKET_NAME')
s3_client = boto3.client(
    's3',
    aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
    aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY')
)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_to_s3(local_file, s3_key):
    try:
        s3_client.upload_file(local_file, S3_BUCKET, s3_key)
        logger.info(f"文件已保存到 S3: {s3_key}")
    except Exception as e:
        logger.error(f"保存到 S3 失败: {str(e)}")
        raise

def load_from_s3(s3_key, local_file):
    try:
        s3_client.download_file(S3_BUCKET, s3_key, local_file)
        logger.info(f"文件已从 S3 下载: {s3_key}")
    except Exception as e:
        logger.error(f"从 S3 下载失败: {str(e)}")
        raise

@app.route('/')
def index():
    try:
        logger.info("访问首页")
        return render_template('index.html')
    except Exception as e:
        logger.error(f"渲染首页时出错: {str(e)}")
        return "服务器错误", 500

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    try:
        # 上传前先备份现有文件
        backup_excel()
        
        if 'file' not in request.files:
            return jsonify({'error': '没有文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400
        
        if file and allowed_file(file.filename):
            try:
                # 使用临时文件进行处理
                temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_upload.xlsx')
                file.save(temp_filepath)
                
                # 验证文件是否为有效的 Excel 文件
                try:
                    wb = openpyxl.load_workbook(temp_filepath)
                    ws = wb.active
                    
                    # 保存到最终位置
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
                    os.makedirs(os.path.dirname(filepath), exist_ok=True)
                    wb.save(filepath)
                    wb.close()
                    
                    logger.info(f"文件保存成功: {filepath}")
                    
                    return jsonify({
                        'message': '文件上传成功（注意：新上传的文件会覆盖现有数据）',
                        'filename': file.filename
                    })
                    
                finally:
                    # 清理临时文件
                    try:
                        if os.path.exists(temp_filepath):
                            os.remove(temp_filepath)
                    except Exception as e:
                        logger.error(f"删除临时文件时出错: {str(e)}")
                        
            except Exception as e:
                logger.error(f"处理Excel文件时出错: {str(e)}")
                return jsonify({'error': '无法处理Excel文件，请确保文件格式正确'}), 400
        
        return jsonify({'error': '不支持的文件类型'}), 400
    except Exception as e:
        logger.error(f"上传文件时出错: {str(e)}")
        return jsonify({'error': '文件上传失败'}), 500

@app.route('/check_number', methods=['POST'])
def check_number():
    try:
        data = request.json
        if not data or 'number' not in data:
            return jsonify({'error': '无效的请求数据'}), 400
            
        number = str(data.get('number')).strip()
        
        # 检查前六位数字
        prefix = number[:6] if len(number) >= 6 else ''
        
        # 定义前缀和对应的列
        prefix_columns = {
            '610423': ('C', 'D'),  # C列数据，D列添加说明
            '610481': ('G', 'H'),  # G列数据，H列添加说明
            '610425': ('K', 'L'),  # K列数据，L列添加说明
            '610424': ('O', 'P')   # O列数据，P列添加说明
        }
        
        if prefix not in prefix_columns:
            return jsonify({'error': '无效的编号前缀'}), 400
            
        # 从 S3 下载最新的文件
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
        load_from_s3('data.xlsx', filepath)
        
        # 处理数据
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        data_col, note_col = prefix_columns[prefix]
        
        # 检查数字是否已存在
        for cell in ws[data_col]:
            if cell.value and str(cell.value).strip() == number:
                wb.close()
                return jsonify({'exists': True, 'message': '该数据已存在'})
        
        # 找到对应列的第一个空单元格
        empty_row = None
        for idx, cell in enumerate(ws[data_col], 1):
            if not cell.value:
                empty_row = idx
                break
        
        if not empty_row:
            empty_row = ws.max_row + 1
                
        # 添加数据
        data_cell = ws[f'{data_col}{empty_row}']
        data_cell.value = number
        
        # 设置红色字体
        from openpyxl.styles import Font
        data_cell.font = Font(color="FF0000")
        
        # 添加说明文字
        ws[f'{note_col}{empty_row}'] = '新添加数据'
        
        # 保存文件
        wb.save(filepath)
        save_to_s3(filepath, 'data.xlsx')
        wb.close()
        
        return jsonify({
            'exists': False, 
            'message': f'新数据已添加到{data_col}列',
            'row': empty_row,
            'column': data_col
        })
            
    except Exception as e:
        logger.error(f"处理请求时出错: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/download_excel')
def download_excel():
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
        # 从 S3 下载最新文件
        load_from_s3('data.xlsx', filepath)
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name='data.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"下载文件时出错: {str(e)}")
        return jsonify({'error': '文件下载失败'}), 500

@app.route('/get_excel_data')
def get_excel_data():
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
        if not os.path.exists(filepath):
            return jsonify({'error': '文件不存在'}), 404
            
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # 获取列名
        columns = [cell.value for cell in ws[1]]
        
        # 获取数据
        data = []
        for row in ws.iter_rows(min_row=2):
            row_data = {}
            for idx, cell in enumerate(row):
                row_data[columns[idx]] = cell.value or ''
            data.append(row_data)
        
        return jsonify({
            'data': data,
            'columns': columns
        })
    except Exception as e:
        logger.error(f"获取Excel数据时出错: {str(e)}")
        return jsonify({'error': '读取数据失败'}), 500

@app.route('/check_file_location')
def check_file_location():
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
        exists = os.path.exists(filepath)
        file_info = {
            'exists': exists,
            'path': filepath,
            'size': os.path.getsize(filepath) if exists else 0,
            'upload_folder': app.config['UPLOAD_FOLDER'],
            'working_dir': os.getcwd()
        }
        return jsonify(file_info)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 添加静态文件缓存控制
@app.after_request
def add_header(response):
    if 'Cache-Control' not in response.headers:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '-1'
    return response

# 添加备份函数
def backup_excel():
    try:
        source = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
        if os.path.exists(source):
            # 创建备份目录
            backup_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'backups')
            os.makedirs(backup_dir, exist_ok=True)
            
            # 生成带时间戳的备份文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_file = os.path.join(backup_dir, f'data_{timestamp}.xlsx')
            
            # 复制文件
            shutil.copy2(source, backup_file)
            logger.info(f'已创建备份: {backup_file}')
            return True
    except Exception as e:
        logger.error(f'创建备份失败: {str(e)}')
        return False

# 添加备份列表接口
@app.route('/list_backups')
def list_backups():
    try:
        backup_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'backups')
        if not os.path.exists(backup_dir):
            return jsonify({'backups': []})
            
        backups = []
        for file in os.listdir(backup_dir):
            if file.startswith('data_') and file.endswith('.xlsx'):
                file_path = os.path.join(backup_dir, file)
                backups.append({
                    'filename': file,
                    'timestamp': file[5:-5],  # 提取时间戳
                    'size': os.path.getsize(file_path)
                })
        
        return jsonify({'backups': sorted(backups, key=lambda x: x['timestamp'], reverse=True)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 添加恢复备份接口
@app.route('/restore_backup/<filename>')
def restore_backup(filename):
    try:
        backup_file = os.path.join(app.config['UPLOAD_FOLDER'], 'backups', filename)
        if not os.path.exists(backup_file):
            return jsonify({'error': '备份文件不存在'}), 404
            
        target = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
        shutil.copy2(backup_file, target)
        
        return jsonify({'message': '恢复成功'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/check_temp_files')
def check_temp_files():
    try:
        # 检查所有可能的目录
        paths_to_check = [
            '/tmp',
            '/opt/render/project/src/uploads',
            '/opt/render/project/src/uploads/backups',
            os.path.join(app.config['UPLOAD_FOLDER']),
            os.getcwd()
        ]
        
        files_found = []
        for path in paths_to_check:
            if os.path.exists(path):
                for file in os.listdir(path):
                    if file.endswith('.xlsx'):
                        file_path = os.path.join(path, file)
                        files_found.append({
                            'path': file_path,
                            'size': os.path.getsize(file_path),
                            'modified': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
                        })
        
        return jsonify({
            'files': files_found,
            'current_dir': os.getcwd(),
            'upload_folder': app.config['UPLOAD_FOLDER']
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/recover_file/<path:file_path>')
def recover_file(file_path):
    try:
        if os.path.exists(file_path):
            # 创建备份目录
            backup_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'backups')
            os.makedirs(backup_dir, exist_ok=True)
            
            # 复制找到的文件
            target = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
            shutil.copy2(file_path, target)
            
            return jsonify({
                'message': '文件已恢复',
                'source': file_path,
                'target': target
            })
        else:
            return jsonify({'error': '文件不存在'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/search_all_files')
def search_all_files():
    try:
        # 定义要搜索的目录列表
        search_paths = [
            '/opt/render/project',  # Render 项目根目录
            '/opt/render/project/src',  # 源代码目录
            '/opt/render/project/src/uploads',  # 上传目录
            '/tmp',  # 临时目录
            os.getcwd(),  # 当前工作目录
            os.path.join(os.getcwd(), 'uploads'),  # uploads 目录
            os.path.join(os.getcwd(), 'backups'),  # backups 目录
        ]
        
        all_files = []
        
        # 递归搜索所有目录
        def scan_directory(path):
            try:
                if os.path.exists(path):
                    for root, dirs, files in os.walk(path):
                        for file in files:
                            if file.endswith(('.xlsx', '.xls')):  # 只搜索 Excel 文件
                                file_path = os.path.join(root, file)
                                all_files.append({
                                    'path': file_path,
                                    'size': os.path.getsize(file_path),
                                    'modified': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S'),
                                    'created': datetime.fromtimestamp(os.path.getctime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
                                })
            except Exception as e:
                logger.error(f"扫描目录失败 {path}: {str(e)}")
        
        # 搜索所有目录
        for path in search_paths:
            scan_directory(path)
            
        # 按修改时间排序
        all_files.sort(key=lambda x: x['modified'], reverse=True)
        
        return jsonify({
            'files': all_files,
            'search_paths': search_paths,
            'total_files': len(all_files)
        })
        
    except Exception as e:
        logger.error(f"搜索文件失败: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(
        host='0.0.0.0',  # 允许所有设备访问
        port=port,
        debug=True
    ) 